import pinyin
from openpyxl import Workbook
import func
from words import *

# Create a new workbook
workbook = Workbook()

# Create a new sheet for each character
sheet = workbook.create_sheet(title='cnyChar')

lst = func.setUpLst(hsk6)

counter = 0
# Loop through rows
for row in range(1, 100, 3):
    # Loop through columns (A to T)
    for col in range(1, 21):
        if counter == len(lst):
            break
        cell1 = sheet.cell(row=row, column=col)
        cell2 = sheet.cell(row=row + 1, column=col)
        # print(cell1.coordinate, cell2.coordinate)

        sheet[str(cell1.coordinate)] = lst[counter]
        sheet[str(cell2.coordinate)] = pinyin.get(lst[counter])
        counter += 1

# Remove the default sheet created initially
workbook.remove(workbook["Sheet"])

# Save the workbook
workbook.save("output6.xlsx")